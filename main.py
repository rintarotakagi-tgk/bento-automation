import os
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NOTION_API_KEY = os.environ["NOTION_API_KEY"]
NOTION_DATABASE_ID = os.environ["NOTION_DATABASE_ID"]
SLACK_BOT_TOKEN = os.environ["SLACK_BOT_TOKEN"]
SLACK_CHANNEL = os.environ["SLACK_CHANNEL"]

TODAY = date.today().isoformat()
TODAY_LABEL = date.today().strftime("%Y年%m月%d日")


def fetch_orders():
    url = f"https://api.notion.com/v1/databases/{NOTION_DATABASE_ID}/query"
    headers = {
        "Authorization": f"Bearer {NOTION_API_KEY}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json",
    }
    payload = {
        "filter": {
            "property": "注文日",
            "date": {"equals": TODAY}
        },
        "sorts": [{"property": "注文者名", "direction": "ascending"}]
    }
    res = requests.post(url, headers=headers, json=payload)
    res.raise_for_status()
    results = res.json().get("results", [])

    orders = []
    for page in results:
        props = page["properties"]

        name_prop = props.get("注文者名", {})
        if name_prop.get("type") == "title":
            name = "".join([t["plain_text"] for t in name_prop.get("title", [])])
        else:
            name = "".join([t["plain_text"] for t in name_prop.get("rich_text", [])])

        item_prop = props.get("注文内容", {})
        item = "".join([t["plain_text"] for t in item_prop.get("rich_text", [])])

        if name:
            orders.append({"name": name, "item": item})

    return orders


def create_excel(orders, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "お弁当注文"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True

    header_fill = PatternFill("solid", start_color="2F5496")
    alt_fill = PatternFill("solid", start_color="DCE6F1")
    white_fill = PatternFill("solid", start_color="FFFFFF")
    gray_fill = PatternFill("solid", start_color="F2F2F2")
    thin = Side(style="thin", color="AAAAAA")
    medium = Side(style="medium", color="2F5496")

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"お弁当注文リスト　{TODAY_LABEL}"
    title_cell.font = Font(name="メイリオ", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = gray_fill
    title_cell.border = Border(left=medium, right=medium, top=medium, bottom=thin)
    ws["D1"].border = Border(left=thin, right=medium, top=medium, bottom=thin)
    ws.row_dimensions[1].height = 36

    headers = ["No.", "注文者名", "注文内容", "✓"]
    ws.append(headers)
    header_row = 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = h
        cell.font = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=medium if col == 1 else thin,
            right=medium if col == 4 else thin,
            top=thin, bottom=thin
        )
    ws.row_dimensions[header_row].height = 24

    last_data_row = header_row
    for i, order in enumerate(orders, 1):
        row_num = header_row + i
        last_data_row = row_num
        fill = alt_fill if i % 2 == 0 else white_fill
        for col, (val, align) in enumerate(zip([i, order["name"], order["item"], ""], ["center", "left", "left", "center"]), 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = val
            cell.font = Font(name="メイリオ", size=11)
            cell.fill = fill
            cell.border = Border(
                left=medium if col == 1 else thin,
                right=medium if col == 4 else thin,
                top=thin, bottom=thin
            )
            cell.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align == "left" else 0, wrap_text=True if col == 3 else False)
        ws.row_dimensions[row_num].height = 25

    total_row = last_data_row + 1
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"].value = f"合計：{len(orders)} 名"
    ws[f"A{total_row}"].font = Font(name="メイリオ", bold=True, size=11)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{total_row}"].fill = gray_fill
    ws[f"A{total_row}"].border = Border(left=medium, right=thin, top=thin, bottom=medium)
    ws.cell(row=total_row, column=4).fill = gray_fill
    ws.cell(row=total_row, column=4).border = Border(left=thin, right=medium, top=thin, bottom=medium)
    ws.row_dimensions[total_row].height = 22

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 7

    ws.print_area = f"A1:D{total_row}"

    wb.save(filepath)
    print(f"Excel作成完了: {filepath}")


def upload_to_slack(filepath):
    filename = os.path.basename(filepath)
    headers = {"Authorization": f"Bearer {SLACK_BOT_TOKEN}"}

    res = requests.get(
        "https://slack.com/api/files.getUploadURLExternal",
        headers=headers,
        params={"filename": filename, "length": os.path.getsize(filepath)}
    )
    res.raise_for_status()
    data = res.json()
    if not data.get("ok"):
        raise RuntimeError(f"URL取得失敗: {data}")

    upload_url = data["upload_url"]
    file_id = data["file_id"]

    with open(filepath, "rb") as f:
        res = requests.post(upload_url, files={"file": f})
    res.raise_for_status()

    res = requests.post(
        "https://slack.com/api/files.completeUploadExternal",
        headers={**headers, "Content-Type": "application/json"},
        json={
            "files": [{"id": file_id}],
            "channel_id": SLACK_CHANNEL,
            "initial_comment": f"🍱 {TODAY_LABEL} のお弁当注文リストです。印刷してご使用ください。"
        }
    )
    res.raise_for_status()
    data = res.json()
    if not data.get("ok"):
        raise RuntimeError(f"投稿失敗: {data}")

    print("Slackへのファイル投稿完了")


def main():
    print(f"=== お弁当注文Slack通知 ({TODAY}) ===")

    orders = fetch_orders()
    print(f"注文件数: {len(orders)} 件")

    if not orders:
        requests.post(
            "https://slack.com/api/chat.postMessage",
            headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}", "Content-Type": "application/json"},
            json={"channel": SLACK_CHANNEL, "text": f"🍱 {TODAY_LABEL} の注文はありませんでした。"}
        )
        print("注文なし通知を送信しました。")
        return

    excel_path = f"/tmp/bento_{TODAY}.xlsx"
    create_excel(orders, excel_path)
    upload_to_slack(excel_path)

    print("=== 完了 ===")


if __name__ == "__main__":
    main()
